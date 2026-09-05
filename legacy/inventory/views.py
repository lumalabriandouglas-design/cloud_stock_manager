import csv
import json
import os
from datetime import timedelta
from decimal import Decimal, InvalidOperation
import io

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from google import genai
from google.genai import types
from PIL import Image

from .models import (
    ActivityLog,
    Category,
    Company,
    Item,
    Sale,
    StaffInvite,
    StockIn,
    Subscription,
    UserProfile,
)


def get_profile(request):
    if not hasattr(request.user, "profile"):
        return None
    return request.user.profile


def get_user_company(request):
    profile = get_profile(request)
    return profile.company if profile else None


def is_superuser(user):
    return user.is_authenticated and user.is_superuser


def require_perm(perm_name):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            profile = get_profile(request)
            if profile is None:
                return redirect("setup_company")
            if not profile.has_perm(perm_name):
                messages.error(request, "You do not have permission for this action.")
                return redirect("dashboard")
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


def require_active_sub(view_func):
    def wrapper(request, *args, **kwargs):
        company = get_user_company(request)
        if company and not company.is_subscription_active():
            messages.error(
                request,
                "Your subscription is inactive. Please renew to continue recording sales and stock.",
            )
            return redirect("billing")
        return view_func(request, *args, **kwargs)
    return wrapper


def perm_context(profile):
    if profile is None:
        return {}
    return {
        "can_manage_stock": profile.has_perm("can_manage_stock"),
        "can_edit_items": profile.has_perm("can_edit_items"),
        "can_view_reports": profile.has_perm("can_view_reports"),
        "can_manage_categories": profile.has_perm("can_manage_categories"),
        "can_manage_team": profile.has_perm("can_manage_team"),
        "is_owner": profile.is_owner,
    }


def log_activity(company, user, action, message):
    ActivityLog.objects.create(company=company, user=user, action=action, message=message)


def notify_low_stock(company, item):
    if not company.low_stock_email_alerts:
        return
    if item.quantity_in_stock > item.reorder_level:
        return
    owners = UserProfile.objects.filter(company=company, role=UserProfile.ROLE_OWNER).select_related("user")
    emails = [p.user.email for p in owners if p.user.email]
    if not emails:
        return
    try:
        send_mail(
            f"[{company.name}] Low stock: {item.name}",
            f"Item: {item.name}\nCurrent stock: {item.quantity_in_stock}\nReorder level: {item.reorder_level}\n",
            None, emails, fail_silently=True,
        )
    except Exception:
        pass


def notify_admin_payment_claim(company, sub):
    admin_email = os.getenv("PAYMENT_NOTIFY_EMAIL", "")
    if not admin_email:
        return
    try:
        send_mail(
            f"[Payment claim] {company.name} - UGX {Subscription.PLAN_AMOUNT_UGX:,}",
            f"Company: {company.name}\nPhone: {sub.payment_phone or '-'}\nTx: {sub.payment_tx_id or '-'}\n",
            None, [admin_email], fail_silently=True,
        )
    except Exception:
        pass


def register(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    if request.method == "POST":
        company_name = request.POST.get("company_name", "").strip()
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        password2 = request.POST.get("password2", "")
        email = request.POST.get("email", "").strip()
        from .features import apply_invite
        pending_invite = StaffInvite.objects.filter(email__iexact=email).first() if email else None
        errors = []
        if not pending_invite and not company_name:
            errors.append("Business name is required.")
        if not username:
            errors.append("Username is required.")
        if len(password) < 6:
            errors.append("Password must be at least 6 characters.")
        if password != password2:
            errors.append("Passwords do not match.")
        if not pending_invite and company_name and Company.objects.filter(name__iexact=company_name).exists():
            errors.append("A business with that name already exists.")
        if User.objects.filter(username__iexact=username).exists():
            errors.append("That username is already taken.")
        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, "inventory/register.html")
        user = User.objects.create_user(username=username, email=email or "", password=password)
        if pending_invite:
            apply_invite(user)
            login(request, user)
            messages.success(request, f"Welcome to {user.profile.company.name}.")
            return redirect("dashboard")
        company = Company.objects.create(name=company_name)
        Subscription.start_trial(company)
        UserProfile.objects.create(
            user=user, company=company, role=UserProfile.ROLE_OWNER,
            can_manage_stock=True, can_edit_items=True, can_view_reports=True,
            can_manage_categories=True, can_manage_team=True,
        )
        login(request, user)
        messages.success(request, f"Welcome! 7-day trial started. Plan UGX {Subscription.PLAN_AMOUNT_UGX:,}/month after trial.")
        return redirect("dashboard")
    return render(request, "inventory/register.html")


@login_required
def setup_company(request):
    if hasattr(request.user, "profile"):
        return redirect("dashboard")
    if request.method == "POST":
        name = request.POST.get("company_name", "").strip()
        if not name:
            messages.error(request, "Please enter a company name.")
            return render(request, "inventory/setup_company.html")
        if Company.objects.filter(name__iexact=name).exists():
            messages.error(request, "A company with that name already exists.")
            return render(request, "inventory/setup_company.html")
        company = Company.objects.create(name=name)
        Subscription.start_trial(company)
        UserProfile.objects.create(
            user=request.user, company=company, role=UserProfile.ROLE_OWNER,
            can_manage_stock=True, can_edit_items=True, can_view_reports=True,
            can_manage_categories=True, can_manage_team=True,
        )
        messages.success(request, f"Company '{company.name}' created. 7-day trial started.")
        return redirect("dashboard")
    return render(request, "inventory/setup_company.html")


@login_required
def dashboard(request):
    profile = get_profile(request)
    if profile is None:
        if request.user.is_superuser:
            return redirect("platform_admin")
        return redirect("setup_company")
    company = profile.company
    all_items = Item.objects.filter(company=company).select_related("category")
    items = all_items.order_by("category__name", "name")
    categories = Category.objects.filter(company=company).order_by("name")
    q = request.GET.get("q", "").strip()
    category_id = request.GET.get("category", "")
    low_only = request.GET.get("low") == "1"
    if q:
        items = items.filter(Q(name__icontains=q) | Q(category__name__icontains=q))
    if category_id:
        items = items.filter(category_id=category_id)
    if low_only:
        items = items.filter(quantity_in_stock__lte=F("reorder_level"))
    recent_activity = ActivityLog.objects.filter(company=company).select_related("user")[:8]
    low_stock_items = all_items.filter(quantity_in_stock__lte=F("reorder_level")).order_by("quantity_in_stock")[:10]
    low_stock_count = all_items.filter(quantity_in_stock__lte=F("reorder_level")).count()
    today = timezone.now().date()
    today_sales_total = (
        Sale.objects.filter(company=company, sales_date__date=today)
        .aggregate(total=Sum(F("quantity_sold") * F("sell_price")))["total"] or 0
    )
    total_inventory_val = sum(i.quantity_in_stock * i.buy_price for i in all_items)
    sub = company.subscription
    sub_active = company.is_subscription_active()
    return render(request, "inventory/dashboard.html", {
        "company": company, "items": items, "categories": categories,
        "recent_activity": recent_activity, "low_stock_items": low_stock_items,
        "low_stock_count": low_stock_count, "total_inventory_val": total_inventory_val,
        "today_sales_total": today_sales_total, "profile": profile,
        "is_platform_admin": request.user.is_superuser, "q": q,
        "selected_category": category_id, "low_only": low_only,
        "subscription": sub, "sub_active": sub_active, **perm_context(profile),
    })


@login_required
def billing(request):
    profile = get_profile(request)
    if profile is None:
        return redirect("setup_company")
    company = profile.company
    sub = company.subscription or Subscription.start_trial(company)
    return render(request, "inventory/billing.html", {
        "company": company, "subscription": sub, "sub_active": sub.is_active,
        "plan_amount": Subscription.PLAN_AMOUNT_UGX, "profile": profile, **perm_context(profile),
    })


@login_required
@require_POST
def claim_payment(request):
    profile = get_profile(request)
    if profile is None or not profile.is_owner:
        messages.error(request, "Only the business owner can submit a payment claim.")
        return redirect("billing")
    company = profile.company
    sub = company.subscription or Subscription.start_trial(company)
    sub.claim_payment(
        phone=request.POST.get("payment_phone", "").strip(),
        tx_id=request.POST.get("payment_tx_id", "").strip(),
        note=request.POST.get("payment_note", "").strip(),
    )
    notify_admin_payment_claim(company, sub)
    messages.success(request, "Payment claim received. We will activate after confirming MoMo.")
    return redirect("billing")


def _norm_header(h):
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def _parse_rows_from_csv(file_obj):
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(raw))
    return [{_norm_header(k): (v or "").strip() for k, v in row.items() if k} for row in reader]


def _parse_rows_from_xlsx(file_obj):
    from openpyxl import load_workbook
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [_norm_header(h) for h in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for values in rows_iter:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = values[i] if i < len(values) else None
            row[h] = "" if val is None else str(val).strip()
        rows.append(row)
    return rows


def _parse_rows_from_pdf(file_obj):
    import pdfplumber
    rows = []
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                if not table or len(table) < 2:
                    continue
                headers = [_norm_header(h) for h in table[0]]
                for data in table[1:]:
                    if not data or all(c is None or str(c).strip() == "" for c in data):
                        continue
                    row = {}
                    for i, h in enumerate(headers):
                        if not h:
                            continue
                        val = data[i] if i < len(data) else None
                        row[h] = "" if val is None else str(val).strip()
                    rows.append(row)
    return rows


def _row_get(row, *keys, default=""):
    for k in keys:
        if k in row and row[k] != "":
            return row[k]
    return default


def _to_decimal(val, default=0):
    if val is None or val == "":
        return Decimal(str(default))
    s = str(val).replace(",", "").replace("UGX", "").strip()
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(str(default))


def _to_int(val, default=0):
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return default


@login_required
@require_perm("can_manage_stock")
@require_active_sub
def import_inventory(request):
    profile = get_profile(request)
    company = profile.company
    if request.method == "POST":
        f = request.FILES.get("import_file")
        if not f:
            messages.error(request, "Please choose a file.")
            return redirect("import_inventory")
        name = (f.name or "").lower()
        try:
            if name.endswith(".csv"):
                rows = _parse_rows_from_csv(f)
            elif name.endswith(".xlsx") or name.endswith(".xls"):
                rows = _parse_rows_from_xlsx(f)
            elif name.endswith(".pdf"):
                rows = _parse_rows_from_pdf(f)
            else:
                messages.error(request, "Unsupported type. Use CSV, Excel (.xlsx), or PDF.")
                return redirect("import_inventory")
        except Exception as e:
            messages.error(request, f"Could not read file: {e}")
            return redirect("import_inventory")
        if not rows:
            messages.error(request, "No data rows found. Need a header row and data.")
            return redirect("import_inventory")
        created = updated = skipped = 0
        for row in rows:
            product = _row_get(row, "name", "product", "item", "product_name", "item_name")
            if not product:
                skipped += 1
                continue
            normalized = " ".join(str(product).split()).title()
            buy = _to_decimal(_row_get(row, "buy_price", "buy", "cost", "cost_price"))
            sell = _to_decimal(_row_get(row, "sell_price", "sell", "price", "selling_price"))
            qty = _to_int(_row_get(row, "quantity", "qty", "stock", "quantity_in_stock"), 0)
            reorder = _to_int(_row_get(row, "reorder_level", "reorder", "min_stock"), 5)
            cat_name = _row_get(row, "category", "cat")
            category = None
            if cat_name:
                category, _ = Category.objects.get_or_create(company=company, name=str(cat_name).strip().title())
            item, is_new = Item.objects.get_or_create(
                company=company, name=normalized,
                defaults={
                    "category": category, "buy_price": buy, "sell_price": sell,
                    "quantity_in_stock": max(0, qty), "reorder_level": reorder if reorder > 0 else 5,
                },
            )
            if is_new:
                created += 1
                if qty > 0:
                    StockIn.objects.create(company=company, item=item, quantity_added=qty)
            else:
                updated += 1
                if qty > 0:
                    item.quantity_in_stock += qty
                    StockIn.objects.create(company=company, item=item, quantity_added=qty)
                if buy > 0:
                    item.buy_price = buy
                if sell > 0:
                    item.sell_price = sell
                if category and not item.category:
                    item.category = category
                if reorder > 0:
                    item.reorder_level = reorder
                item.save()
        log_activity(company, request.user, ActivityLog.ACTION_STOCK_IN, f"Import: {created} new, {updated} updated")
        messages.success(request, f"Import done: {created} created, {updated} updated" + (f", {skipped} skipped." if skipped else "."))
        return redirect("dashboard")
    return render(request, "inventory/import_inventory.html", {"company": company, "profile": profile, **perm_context(profile)})


@login_required
@require_perm("can_view_reports")
def export_inventory_csv(request):
    company = get_user_company(request)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{company.name}_inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(["Name", "Category", "Buy Price", "Sell Price", "Quantity", "Reorder Level"])
    for item in Item.objects.filter(company=company).select_related("category").order_by("name"):
        writer.writerow([item.name, item.category.name if item.category else "", item.buy_price, item.sell_price, item.quantity_in_stock, item.reorder_level])
    return response


@login_required
@require_perm("can_view_reports")
def export_sales_csv(request):
    company = get_user_company(request)
    days = int(request.GET.get("days", 30))
    since = timezone.now() - timezone.timedelta(days=days)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{company.name}_sales.csv"'
    writer = csv.writer(response)
    writer.writerow(["Date", "Product", "Qty", "Unit Price", "Revenue", "Est. Cost", "Est. Profit"])
    for s in Sale.objects.filter(company=company, sales_date__gte=since).select_related("item").order_by("-sales_date"):
        writer.writerow([s.sales_date.strftime("%Y-%m-%d %H:%M"), s.item.name, s.quantity_sold, s.sell_price, s.line_total, s.estimated_cost, s.estimated_profit])
    return response


@login_required
@require_perm("can_manage_team")
def manage_team(request):
    profile = get_profile(request)
    company = profile.company
    members = UserProfile.objects.filter(company=company).select_related("user").order_by("role", "user__username")
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            username = request.POST.get("username", "").strip()
            password = request.POST.get("password", "")
            email = request.POST.get("email", "").strip().lower()
            role = request.POST.get("role", UserProfile.ROLE_STAFF)
            if email and not username:
                if StaffInvite.objects.filter(company=company, email=email).exists():
                    messages.error(request, "That email already has an invite.")
                else:
                    StaffInvite.objects.create(
                        company=company,
                        email=email,
                        role=role,
                        can_manage_stock=request.POST.get("can_manage_stock") == "on" or role == UserProfile.ROLE_OWNER,
                        can_edit_items=request.POST.get("can_edit_items") == "on" or role == UserProfile.ROLE_OWNER,
                        can_view_reports=request.POST.get("can_view_reports") == "on" or role == UserProfile.ROLE_OWNER,
                        can_manage_categories=request.POST.get("can_manage_categories") == "on" or role == UserProfile.ROLE_OWNER,
                        can_manage_team=request.POST.get("can_manage_team") == "on" or role == UserProfile.ROLE_OWNER,
                    )
                    messages.success(request, f"Invite saved for {email}. They join when they sign in or register with that email.")
            elif not username or len(password) < 6:
                messages.error(request, "Username and password (min 6 chars) required — or invite by email only.")
            elif User.objects.filter(username__iexact=username).exists():
                messages.error(request, "Username already taken.")
            else:
                user = User.objects.create_user(username=username, email=email, password=password)
                UserProfile.objects.create(
                    user=user, company=company, role=role,
                    can_manage_stock=request.POST.get("can_manage_stock") == "on" or role == UserProfile.ROLE_OWNER,
                    can_edit_items=request.POST.get("can_edit_items") == "on" or role == UserProfile.ROLE_OWNER,
                    can_view_reports=request.POST.get("can_view_reports") == "on" or role == UserProfile.ROLE_OWNER,
                    can_manage_categories=request.POST.get("can_manage_categories") == "on" or role == UserProfile.ROLE_OWNER,
                    can_manage_team=request.POST.get("can_manage_team") == "on" or role == UserProfile.ROLE_OWNER,
                )
                messages.success(request, f"Added {username}.")
        elif action == "cancel_invite":
            StaffInvite.objects.filter(id=request.POST.get("invite_id"), company=company).delete()
            messages.success(request, "Invite cancelled.")
        elif action == "update_perms":
            member = get_object_or_404(UserProfile, id=request.POST.get("profile_id"), company=company)
            if member.is_owner:
                messages.error(request, "Cannot change Owner permissions.")
            else:
                member.can_manage_stock = request.POST.get("can_manage_stock") == "on"
                member.can_edit_items = request.POST.get("can_edit_items") == "on"
                member.can_view_reports = request.POST.get("can_view_reports") == "on"
                member.can_manage_categories = request.POST.get("can_manage_categories") == "on"
                member.can_manage_team = request.POST.get("can_manage_team") == "on"
                member.save()
                messages.success(request, f"Updated {member.user.username}.")
        elif action == "remove":
            member = get_object_or_404(UserProfile, id=request.POST.get("profile_id"), company=company)
            if member.user == request.user:
                messages.error(request, "You cannot remove yourself.")
            elif member.is_owner and UserProfile.objects.filter(company=company, role=UserProfile.ROLE_OWNER).count() <= 1:
                messages.error(request, "Cannot remove the last owner.")
            else:
                uname = member.user.username
                member.user.delete()
                messages.success(request, f"Removed {uname}.")
        return redirect("manage_team")
    invites = StaffInvite.objects.filter(company=company).order_by("-created_at")
    return render(request, "inventory/manage_team.html", {"company": company, "members": members, "invites": invites, "profile": profile, **perm_context(profile)})


@login_required
@require_perm("can_manage_categories")
def manage_categories(request):
    profile = get_profile(request)
    company = profile.company
    categories = Category.objects.filter(company=company).order_by("name")
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            name = request.POST.get("name", "").strip()
            if not name:
                messages.error(request, "Name required.")
            elif Category.objects.filter(company=company, name__iexact=name).exists():
                messages.error(request, "Category already exists.")
            else:
                Category.objects.create(company=company, name=name.title())
                messages.success(request, "Category added.")
        elif action == "edit":
            cat = get_object_or_404(Category, id=request.POST.get("category_id"), company=company)
            name = request.POST.get("name", "").strip()
            if name and not Category.objects.filter(company=company, name__iexact=name).exclude(id=cat.id).exists():
                cat.name = name.title()
                cat.save()
                messages.success(request, "Updated.")
            else:
                messages.error(request, "Invalid or duplicate name.")
        elif action == "delete":
            get_object_or_404(Category, id=request.POST.get("category_id"), company=company).delete()
            messages.success(request, "Deleted.")
        return redirect("manage_categories")
    return render(request, "inventory/manage_categories.html", {"company": company, "categories": categories, "profile": profile, **perm_context(profile)})


@login_required
@require_perm("can_edit_items")
def edit_item(request, item_id):
    profile = get_profile(request)
    company = profile.company
    item = get_object_or_404(Item, id=item_id, company=company)
    categories = Category.objects.filter(company=company).order_by("name")
    if request.method == "POST":
        if not company.is_subscription_active():
            messages.error(request, "Subscription inactive. Renew to edit.")
            return redirect("billing")
        name = request.POST.get("name", "").strip()
        category_id = request.POST.get("category_id")
        buy_price = request.POST.get("buy_price") or 0
        sell_price = request.POST.get("sell_price") or 0
        reorder_level = request.POST.get("reorder_level") or 5
        qty_adjust = request.POST.get("qty_adjust")
        if not name:
            messages.error(request, "Name required.")
        elif Item.objects.filter(company=company, name__iexact=name).exclude(id=item.id).exists():
            messages.error(request, "Name already used.")
        else:
            item.name = name
            item.category = Category.objects.filter(id=category_id, company=company).first() if category_id else None
            item.buy_price = buy_price
            item.sell_price = sell_price
            item.reorder_level = int(reorder_level)
            if qty_adjust and profile.has_perm("can_manage_stock"):
                try:
                    delta = int(qty_adjust)
                    item.quantity_in_stock = max(0, item.quantity_in_stock + delta)
                    if delta > 0:
                        StockIn.objects.create(company=company, item=item, quantity_added=delta)
                except ValueError:
                    pass
            item.save()
            log_activity(company, request.user, ActivityLog.ACTION_ITEM_EDIT, f"Updated {item.name}")
            messages.success(request, f"'{item.name}' updated.")
            return redirect("dashboard")
    return render(request, "inventory/edit_item.html", {"company": company, "item": item, "categories": categories, "profile": profile, **perm_context(profile)})


@login_required
@require_perm("can_view_reports")
def sales_report(request):
    profile = get_profile(request)
    company = profile.company
    days = int(request.GET.get("days", 30))
    since = timezone.now() - timezone.timedelta(days=days)
    sales = Sale.objects.filter(company=company, sales_date__gte=since).select_related("item").order_by("-sales_date")
    total_revenue = sum(s.line_total for s in sales)
    total_cost = sum(s.estimated_cost for s in sales)
    return render(request, "inventory/sales_report.html", {
        "company": company, "sales": sales, "days": days,
        "total_revenue": total_revenue, "total_cost": total_cost,
        "total_profit": total_revenue - total_cost, "total_qty": sum(s.quantity_sold for s in sales),
        "profile": profile, "is_platform_admin": request.user.is_superuser, **perm_context(profile),
    })


@login_required
@require_perm("can_manage_stock")
@require_active_sub
def record_sale(request):
    if request.method == "POST":
        company = get_user_company(request)
        item_id = request.POST.get("item_id")
        quantity = int(request.POST.get("quantity", 1))
        custom_price = request.POST.get("sell_price")
        if not item_id:
            messages.error(request, "Select an item.")
            return redirect("sell")
        item = get_object_or_404(Item, id=item_id, company=company)
        sell_price = custom_price if custom_price else item.sell_price
        if item.quantity_in_stock < quantity:
            messages.error(request, f"Only {item.quantity_in_stock} left of {item.name}.")
            return redirect("sell")
        Sale.objects.create(company=company, item=item, quantity_sold=quantity, sell_price=sell_price)
        item.quantity_in_stock -= quantity
        item.save()
        log_activity(company, request.user, ActivityLog.ACTION_SALE, f"Sold {quantity}x {item.name}")
        notify_low_stock(company, item)
        messages.success(request, f"Sold {quantity} x {item.name}.")
    return redirect("sell")


@login_required
@require_perm("can_manage_stock")
@require_active_sub
def record_stock_in(request):
    if request.method == "POST":
        company = get_user_company(request)
        item_name = request.POST.get("item_name", "").strip()
        category_id = request.POST.get("category_id")
        buy_price = request.POST.get("buy_price") or 0
        sell_price = request.POST.get("sell_price") or 0
        quantity = int(request.POST.get("quantity", 0))
        if not item_name or quantity <= 0:
            messages.error(request, "Name and positive quantity required.")
            return redirect("dashboard")
        normalized = " ".join(item_name.split()).title()
        category = Category.objects.filter(id=category_id, company=company).first() if category_id else None
        item, created = Item.objects.get_or_create(
            company=company, name=normalized,
            defaults={"category": category, "buy_price": buy_price, "sell_price": sell_price, "quantity_in_stock": 0},
        )
        item.quantity_in_stock += quantity
        if buy_price:
            item.buy_price = buy_price
        if sell_price:
            item.sell_price = sell_price
        if category and not item.category:
            item.category = category
        item.save()
        StockIn.objects.create(company=company, item=item, quantity_added=quantity)
        log_activity(company, request.user, ActivityLog.ACTION_STOCK_IN, f"+{quantity} {item.name}")
        messages.success(request, f"{'Created' if created else 'Restocked'} {item.name}: +{quantity}")
    return redirect("dashboard")


@login_required
@require_perm("can_manage_stock")
@require_active_sub
def scan_ledger(request):
    if request.method == "POST" and request.FILES.get("ledger_photo"):
        company = get_user_company(request)
        photo = request.FILES["ledger_photo"]
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            messages.error(request, "GEMINI_API_KEY not set.")
            return redirect("dashboard")
        try:
            img = Image.open(photo)
            client = genai.Client(api_key=api_key)
            prompt = 'Analyze this handwritten stock ledger. Return strict JSON list: name, buy_price, sell_price, quantity.'
            response = client.models.generate_content(
                model="gemini-2.0-flash", contents=[prompt, img],
                config=types.GenerateContentConfig(response_mime_type="application/json"),
            )
            items_data = json.loads(response.text)
            added = 0
            for entry in items_data:
                name = str(entry.get("name", "")).strip()
                if not name:
                    continue
                normalized = " ".join(name.split()).title()
                buy = float(entry.get("buy_price") or 0)
                sell = float(entry.get("sell_price") or 0)
                qty = int(entry.get("quantity") or 1)
                item, _ = Item.objects.get_or_create(company=company, name=normalized, defaults={"buy_price": buy, "sell_price": sell, "quantity_in_stock": 0})
                item.quantity_in_stock += qty
                if buy > 0:
                    item.buy_price = buy
                if sell > 0:
                    item.sell_price = sell
                item.save()
                StockIn.objects.create(company=company, item=item, quantity_added=qty)
                added += 1
            if added:
                log_activity(company, request.user, ActivityLog.ACTION_STOCK_IN, f"AI scan imported {added} item(s)")
                messages.success(request, f"Imported {added} item(s).")
            else:
                messages.warning(request, "No products extracted.")
        except Exception as e:
            messages.error(request, f"Scan error: {e}")
    return redirect("dashboard")


@login_required
@user_passes_test(is_superuser)
def platform_admin(request):
    companies = Company.objects.all().order_by("-created_at")
    pending_subs = Subscription.objects.filter(status=Subscription.STATUS_PENDING).select_related("company").order_by("-payment_claimed_at")
    all_subs = list(Subscription.objects.select_related("company"))
    stats = {
        "companies": companies.count(),
        "active": sum(1 for s in all_subs if s.status == Subscription.STATUS_ACTIVE and s.is_active),
        "trial": sum(1 for s in all_subs if s.status == Subscription.STATUS_TRIAL and s.is_active),
        "pending": pending_subs.count(),
    }
    company_rows = [{"company": c, "sub": getattr(c, "sub", None), "active": c.is_subscription_active(), "users": c.users.count()} for c in companies]
    cutoff = timezone.now() - timezone.timedelta(minutes=5)
    online_users = UserProfile.objects.filter(last_seen__gte=cutoff).select_related("user", "company").order_by("-last_seen")
    return render(request, "inventory/platform_admin.html", {"stats": stats, "pending_subs": pending_subs, "company_rows": company_rows, "online_users": online_users})


@login_required
@user_passes_test(is_superuser)
def platform_create_business(request):
    if request.method == "POST":
        company_name = request.POST.get("company_name", "").strip()
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        email = request.POST.get("email", "").strip()
        errors = []
        if not company_name:
            errors.append("Company name required.")
        if not username:
            errors.append("Username required.")
        if not password or len(password) < 6:
            errors.append("Password min 6 characters.")
        if Company.objects.filter(name__iexact=company_name).exists():
            errors.append("Company name taken.")
        if User.objects.filter(username__iexact=username).exists():
            errors.append("Username taken.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            company = Company.objects.create(name=company_name)
            Subscription.start_trial(company)
            user = User.objects.create_user(username=username, email=email or "", password=password)
            UserProfile.objects.create(user=user, company=company, role=UserProfile.ROLE_OWNER, can_manage_stock=True, can_edit_items=True, can_view_reports=True, can_manage_categories=True, can_manage_team=True)
            messages.success(request, f"Created '{company.name}' - owner: {username}")
            return redirect("platform_admin")
    companies = Company.objects.all().order_by("-created_at")
    pending = Subscription.objects.filter(status=Subscription.STATUS_PENDING).select_related("company").order_by("-payment_claimed_at")
    return render(request, "inventory/platform_create_business.html", {"companies": companies, "pending_subs": pending})


@login_required
@user_passes_test(is_superuser)
def platform_activate_sub(request, sub_id):
    sub = get_object_or_404(Subscription, id=sub_id)
    sub.activate_for_month()
    messages.success(request, f"Activated {sub.company.name} for 30 days.")
    return redirect("platform_admin")


@login_required
@user_passes_test(is_superuser)
def platform_suspend_sub(request, sub_id):
    sub = get_object_or_404(Subscription, id=sub_id)
    sub.mark_suspended()
    messages.success(request, f"Suspended {sub.company.name}.")
    return redirect("platform_admin")


@login_required
@user_passes_test(is_superuser)
def platform_extend_sub(request, sub_id):
    sub = get_object_or_404(Subscription, id=sub_id)
    now = timezone.now()
    if sub.current_period_end and sub.current_period_end > now:
        sub.current_period_end = sub.current_period_end + timedelta(days=30)
    else:
        sub.current_period_end = now + timedelta(days=30)
    sub.status = Subscription.STATUS_ACTIVE
    sub.last_payment_at = now
    sub.save()
    messages.success(request, f"Extended {sub.company.name} by 30 days (until {sub.current_period_end.date()}).")
    return redirect("platform_admin")
